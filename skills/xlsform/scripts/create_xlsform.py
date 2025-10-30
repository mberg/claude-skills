#!/usr/bin/env python3
# ABOUTME: Helper script to create XLSForm Excel files from CSV/TSV data
# ABOUTME: Generates properly structured .xlsx files from survey and choices data

import sys
import csv
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("\nInstall with uv:")
    print("  uv pip install openpyxl")
    print("\nOr with pip:")
    print("  pip install openpyxl")
    sys.exit(1)


def create_xlsform(survey_data, choices_data=None, settings_data=None, output_path='form.xlsx'):
    """
    Create XLSForm Excel file from data.

    Args:
        survey_data: List of dicts with survey questions
        choices_data: List of dicts with choice options (optional)
        settings_data: Dict with form settings (optional)
        output_path: Output .xlsx file path

    Returns:
        Path to created file
    """
    workbook = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Create survey worksheet
    survey_sheet = workbook.create_sheet('survey')
    if survey_data:
        # Get headers from first row
        headers = list(survey_data[0].keys())
        survey_sheet.append(headers)

        # Add data rows
        for row in survey_data:
            survey_sheet.append([row.get(h, '') for h in headers])

    # Create choices worksheet if provided
    if choices_data:
        choices_sheet = workbook.create_sheet('choices')
        headers = list(choices_data[0].keys())
        choices_sheet.append(headers)

        for row in choices_data:
            choices_sheet.append([row.get(h, '') for h in headers])

    # Create settings worksheet if provided
    if settings_data:
        settings_sheet = workbook.create_sheet('settings')
        headers = list(settings_data.keys())
        settings_sheet.append(headers)
        settings_sheet.append([settings_data.get(h, '') for h in headers])

    # Save workbook
    workbook.save(output_path)
    return output_path


def load_csv(csv_path):
    """Load CSV/TSV file into list of dicts."""
    data = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        # Auto-detect delimiter
        sample = f.read(1024)
        f.seek(0)
        delimiter = '\t' if '\t' in sample else ','

        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            # Remove empty values
            cleaned = {k: v for k, v in row.items() if v}
            if cleaned:  # Skip empty rows
                data.append(cleaned)

    return data


def main():
    if len(sys.argv) < 2:
        print("Usage: python create_xlsform.py <survey.csv> [choices.csv] [-o output.xlsx]")
        print("\nCreate XLSForm Excel file from CSV/TSV data.")
        print("\nArguments:")
        print("  survey.csv     Survey worksheet data (required)")
        print("  choices.csv    Choices worksheet data (optional)")
        print("  -o FILE        Output file path (default: form.xlsx)")
        print("\nExample:")
        print("  python create_xlsform.py survey.csv choices.csv -o myform.xlsx")
        sys.exit(1)

    survey_path = sys.argv[1]
    choices_path = None
    output_path = 'form.xlsx'

    # Parse arguments
    i = 2
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == '-o' and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]
            i += 2
        elif not arg.startswith('-'):
            choices_path = arg
            i += 1
        else:
            i += 1

    # Validate inputs
    if not Path(survey_path).exists():
        print(f"ERROR: Survey file not found: {survey_path}")
        sys.exit(1)

    if choices_path and not Path(choices_path).exists():
        print(f"ERROR: Choices file not found: {choices_path}")
        sys.exit(1)

    # Load data
    print(f"Loading survey from: {survey_path}")
    survey_data = load_csv(survey_path)

    choices_data = None
    if choices_path:
        print(f"Loading choices from: {choices_path}")
        choices_data = load_csv(choices_path)

    # Create XLSForm
    print(f"Creating XLSForm: {output_path}")
    result = create_xlsform(survey_data, choices_data, output_path=output_path)

    print(f"✓ XLSForm created successfully: {result}")
    print(f"\nNext steps:")
    print(f"  1. Validate: python scripts/validate_xlsform.py {result}")
    print(f"  2. Convert: python scripts/convert_to_xform.py {result}")

    return 0


if __name__ == '__main__':
    sys.exit(main())
